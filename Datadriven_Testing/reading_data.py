#Program for reading data from the Excel...
import openpyxl

file="D:\\python_selenium\\pythonProject\\Data\\DemoSheet.xlsx"
book=openpyxl.load_workbook(file)
sheet=book["Sheet1"]
#We use active method when we don't have more than 1 sheet in our Excel file
# sheet=book.active     #In simple term active means single sheet.

rows=sheet.max_row        #Count number of rows in Excel sheet    =5
columns=sheet.max_column  #Count number of columns in Excel sheet  =3

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='     ')
    print()