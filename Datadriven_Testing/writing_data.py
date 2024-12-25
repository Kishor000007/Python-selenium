# Program for writing data into Excel

import openpyxl

file="D:\\python_selenium\\pythonProject\\Data\\DemoSheet.xlsx"
book=openpyxl.load_workbook(file)
sheet=book["Sheet2"]

#It will write same data in all rows and columns
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Welcome"
#
# book.save(file)      #We have to write save command save the file after writing into Excel.

#To write different data in rows and columns
sheet.cell(1,1).value="SID"
sheet.cell(1,2).value="SNAME"
sheet.cell(1,3).value="COURSE"

sheet.cell(2,1).value="1"
sheet.cell(2,2).value="Allen"
sheet.cell(2,3).value="Python"

sheet.cell(3,1).value="2"
sheet.cell(3,2).value="Scott"
sheet.cell(3,3).value="Java"

sheet.cell(4,1).value="3"
sheet.cell(4,2).value="Smith"
sheet.cell(4,3).value="HTML"

book.save(file)