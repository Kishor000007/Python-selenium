import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rowNum,colNum):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetName]
    return sheet.cell(rowNum,colNum).value

def writeData(file,sheetName,rowNum,colNum,data):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetName]
    sheet.cell(rowNum,colNum).value = data
    book.save(file)

def fillGreenColor(file,sheetName,rowNum,colNum):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetName]
    greenfill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rowNum,colNum).fill=greenfill
    book.save(file)

def fillRedColor(file,sheetName,rowNum,colNum):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetName]
    redfill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rowNum,colNum).fill=redfill
    book.save(file)